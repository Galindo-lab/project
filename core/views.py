import openpyxl
import csv

from urllib.parse import urlencode
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django import forms
from django.http import HttpResponse
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.contrib import messages
from django.db.models import Q
from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone
from django.shortcuts import get_object_or_404
from django.contrib import messages
from django.urls import reverse
from django.views.generic.edit import UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import get_object_or_404, redirect, render
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.urls import reverse, reverse_lazy
from django.views import View
from django.views.generic import ListView, CreateView, FormView, UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.mixins import UserPassesTestMixin
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string

from project import settings


from .forms import (
    ArchiveProjectForm,
    CreateResourceForm,
    CreateTaskForm,
    PasswordChangeCustomForm,
    ProjectEditForm,
    RegisterForm,
    TaskEditForm,
    EmprendedorDescripcionForm,
    UserEditWithProfileForm,
    UserProfileForm,
)

from .models import (
    Collaborator, 
    Goal, 
    Project, 
    Resource, 
    Task, 
    Collaborator
)

from .generators import GeminiGenerator



class ProjectAccessMixin(UserPassesTestMixin):
    """
    Allows access if the user is staff, owner or a collaborator of the project.
    Requires the view to have 'project_pk' or 'pk' in self.kwargs.
    """

    def test_func(self):
        user = self.request.user
        if user.is_staff:
            return True

        project_pk = self.kwargs.get("project_pk") or self.kwargs.get("pk")
        if not project_pk:
            return False

        project = get_object_or_404(Project, pk=project_pk)
        # Permitir acceso si es dueño o colaborador
        if project.owner == user:
            return True
        return Collaborator.objects.filter(user=user, project=project).exists()





# Usuarios 
@login_required
def editar_colaborador(request, project_id, collaborator_id):
    collaborator = get_object_or_404(
        Collaborator, id=collaborator_id, project_id=project_id
    )
    if request.method == "POST":
        role = request.POST.get("role")
        if role in dict(Collaborator.Permissions.choices):
            collaborator.role = role
            collaborator.save()
            messages.success(request, "Permiso actualizado correctamente.")
        else:
            messages.error(request, "Permiso no válido.")
    return redirect("details_project", pk=project_id)


class UserListView(ListView):
    model = User
    template_name = "view/userList/main.html"
    context_object_name = "users"
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get("search", "").strip()
        status = self.request.GET.get("status", "")

        if search_query:
            queryset = queryset.filter(
                Q(username__icontains=search_query)
                | Q(email__icontains=search_query)
                | Q(first_name__icontains=search_query)
                | Q(last_name__icontains=search_query)
            )
        if status == "activo":
            queryset = queryset.filter(is_active=True)
        elif status == "suspendido":
            queryset = queryset.filter(is_active=False)
        elif status == "eliminado":
            queryset = queryset.filter(is_active=False)  # Ajusta si tienes un campo especial para "eliminado"
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["current_status"] = self.request.GET.get("status", "")
        return context


@login_required
def user_delete(request, pk):
    user = get_object_or_404(User, pk=pk)
    if request.method == "POST":
        user.delete()
        messages.success(request, "Usuario eliminado correctamente.")
        return redirect("user_list")
    return render(request, "view/userList/modal/deleteUser.html", {"user": user})





@login_required
def index(request):
    """
    Render the index page.
    """
    return render(request, "layout/app/main.html")


class UserEditView(UpdateView):
    model = User
    form_class = UserEditWithProfileForm
    template_name = "view/userList/editUser.html"
    success_url = reverse_lazy("user_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["projects"] = Project.objects.filter(owner=self.object)
        return context


def register(request):
    if request.method == "POST":
        form = RegisterForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("home")
    else:
        form = RegisterForm()

    return render(
        request,
        "view/register.html",
        {"form": form},
    )


class InviteCollaboratorView(ProjectAccessMixin, LoginRequiredMixin, View):
    def post(self, request, project_id, *args, **kwargs):
        email = request.POST.get("email")
        project = get_object_or_404(Project, id=project_id)

        if self._is_basic_account_limit_reached(request.user, project):
            messages.error(
                request,
                "Con una cuenta básica solo puedes agregar un colaborador a tu proyecto.",
            )
            return redirect("details_project", pk=project.id)

        user = self._get_user_by_email(email)
        if user:
            if self._is_already_collaborator(user, project):
                messages.info(request, f"{email} ya es colaborador de este proyecto.")
            else:
                self._add_collaborator(user, project, email, request)
        else:
            self._send_invitation_email(email, project, request)

        return redirect("details_project", pk=project.id)

    def _is_basic_account_limit_reached(self, user, project):
        return (
            user == project.owner
            and getattr(user, "account_type", "basic") == "basic"
            and Collaborator.objects.filter(project=project)
                .exclude(user=project.owner)
                .count() >= 1
        )

    def _get_user_by_email(self, email):
        try:
            return User.objects.get(email=email)
        except User.DoesNotExist:
            return None

    def _is_already_collaborator(self, user, project):
        return Collaborator.objects.filter(user=user, project=project).exists()

    def _add_collaborator(self, user, project, email, request):
        Collaborator.objects.create(
            user=user,
            project=project,
            role="Colaborador",
            invitation_date=timezone.now(),
        )
        subject = "Invitación a colaborar en un proyecto"
        message = render_to_string(
            "emails/colaborador_agregado.txt",
            {"project": project}
        )
        send_mail(
            subject,
            message,
            "no-reply@tusitio.com",
            [email],
            fail_silently=False,
        )
        messages.success(
            request,
            f"{email} ha sido agregado como colaborador y notificado por correo.",
        )

    def _send_invitation_email(self, email, project, request):
        subject = "Invitación a colaborar en un proyecto"
        message = render_to_string(
            "emails/invitacion_colaborador.txt",
            {"project": project}
        )
        send_mail(
            subject,
            message,
            "no-reply@tusitio.com",
            [email],
            fail_silently=False,
        )
        messages.success(
            request,
            f"Se ha enviado una invitación a {email}. Cuando se registre, podrá ser agregado como colaborador.",
        )


class UserProfileView(LoginRequiredMixin, View):
    template_name = "view/profile_edit.html"

    def get(self, request):
        profile_form = UserProfileForm(instance=request.user)
        password_form = PasswordChangeCustomForm()
        return render(request, self.template_name, {
            "profile_form": profile_form,
            "password_form": password_form,
        })

    def post(self, request):
        profile_form = UserProfileForm(request.POST, instance=request.user)
        password_form = PasswordChangeCustomForm(request.POST)

        if "update_profile" in request.POST:
            if profile_form.is_valid():
                profile_form.save()
                messages.success(request, "Perfil actualizado correctamente.")
                return redirect("profile_edit")
        elif "change_password" in request.POST:
            if password_form.is_valid():
                old_password = password_form.cleaned_data["old_password"]
                new_password1 = password_form.cleaned_data["new_password1"]
                new_password2 = password_form.cleaned_data["new_password2"]
                if not request.user.check_password(old_password):
                    messages.error(request, "La contraseña actual es incorrecta.")
                elif new_password1 != new_password2:
                    messages.error(request, "Las nuevas contraseñas no coinciden.")
                else:
                    request.user.set_password(new_password1)
                    request.user.save()
                    update_session_auth_hash(request, request.user)
                    messages.success(request, "Contraseña cambiada correctamente.")
                    return redirect("profile_edit")
        return render(request, self.template_name, {
            "profile_form": profile_form,
            "password_form": password_form,
        })


class CollaboratorDeleteView(ProjectAccessMixin, LoginRequiredMixin, View):
    def post(self, request, project_id, *args, **kwargs):
        collaborator_id = request.POST.get("collaborator_id")
        if not collaborator_id:
            messages.error(request, "No se especificó el colaborador a eliminar.")
            return redirect("details_project", pk=project_id)

        project = get_object_or_404(Project, id=project_id)
        collaborator = get_object_or_404(Collaborator, id=collaborator_id, project=project)

        if collaborator.user == project.owner:
            messages.error(request, "No puedes eliminar al propietario del proyecto.")
        else:
            collaborator.delete()
            messages.success(request, "Colaborador eliminado correctamente.")

        return redirect("details_project", pk=project_id)


# Vistas de Recursos
class ResourceEditView(LoginRequiredMixin, View):
    def post(self, request, project_pk, *args, **kwargs):
        resource_id = request.POST.get("resource_id")
        resource = get_object_or_404(Resource, pk=resource_id, project__pk=project_pk)

        resource.name = request.POST.get("name")
        resource.type = request.POST.get("type")
        resource.cost_per_hour = request.POST.get("cost_per_hour")
        resource.save()

        messages.success(request, "Recurso modificado correctamente!")
        return redirect(reverse("resources", kwargs={"project_pk": project_pk}))


class ResourceDeleteView(LoginRequiredMixin, View):
    def post(self, request, project_pk, *args, **kwargs):
        resource_id = request.POST.get("resource_id")
        project = get_object_or_404(Project, pk=project_pk)
        resource = get_object_or_404(Resource, pk=resource_id, project=project)

        resource.delete()
        messages.success(request, "Recurso eliminado correctamente!")
        return redirect(reverse("resources", kwargs={"project_pk": project.pk}))


class ResourceCreateView(LoginRequiredMixin, View):
    def post(self, request, project_pk, *args, **kwargs):
        project = get_object_or_404(Project, pk=project_pk)

        # URL de redirección
        redirect_url = reverse("resources", kwargs={"project_pk": project.pk})

        resource_form = CreateResourceForm(request.POST)
        if resource_form.is_valid():
            resource = resource_form.save(commit=False)
            resource.project = project
            resource.save()
            messages.success(request, "Recurso creado correctamente!")
            return redirect(redirect_url)

        messages.error(
            request,
            "Error al crear el recurso. Por favor, verifica los datos ingresados.",
        )
        return redirect(redirect_url)


class ResourcesListView(LoginRequiredMixin, ListView):
    model = Resource
    template_name = "view/resourcesList/main.html"
    context_object_name = "resources"
    paginate_by = 10

    def get_queryset(self):
        project = get_object_or_404(Project, pk=self.kwargs["project_pk"])
        search_query = self.request.GET.get("search", "").strip()

        if search_query:
            return project.resource_set.filter(Q(name__icontains=search_query))

        return project.resource_set.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project = get_object_or_404(Project, pk=self.kwargs["project_pk"])
        context["project"] = project
        context["resource_type_choices"] = (
            Resource.ResourceType.choices
        )  # Agregar las opciones de tipo de recurso
        return context


class ResourceGenerateAIView(LoginRequiredMixin, View):
    def get(self, request, project_pk, *args, **kwargs):
        project = get_object_or_404(Project, pk=project_pk)
        gg = GeminiGenerator()
        resource_data = gg.generate_resource(project, project.description)
        # Crea el recurso en la base de datos
        Resource.objects.create(
            name=resource_data["name"],
            type=resource_data["type"],
            cost_per_hour=resource_data["cost_per_hour"],
            project=project,
        )
        messages.success(request, "Recurso generado correctamente con IA.")
        return redirect(reverse("resources", kwargs={"project_pk": project.pk}))

# Vistas de Tareas
class TaskDetailView(LoginRequiredMixin, UpdateView):
    model = Task
    form_class = TaskEditForm
    template_name = "view/taskDetails.html"

    def get_object(self):
        task_id = self.kwargs.get("task_pk")
        return get_object_or_404(Task, pk=task_id)

    def get_success_url(self):
        task = self.get_object()
        return reverse("details_task", kwargs={"task_pk": task.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        task = self.get_object()
        goal = get_object_or_404(Goal, pk=task.goal.pk)
        project = get_object_or_404(Project, pk=goal.project.pk)
        context["project"] = project
        # Limitar los recursos del formulario solo a los del proyecto
        context["form"].fields["resources"].queryset = project.resource_set.all()
        return context

    def form_valid(self, form):
        messages.success(self.request, "¡Tarea actualizada correctamente!")
        return super().form_valid(form)

    def form_invalid(self, form):
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(
                    self.request, f"Error en {form.fields[field].label}: {error}"
                )
        return super().form_invalid(form)


class TaskCreateView(LoginRequiredMixin, View):
    def post(self, request, goal_pk, *args, **kwargs):
        goal = get_object_or_404(Goal, pk=goal_pk)
        project = get_object_or_404(Project, pk=goal.project.pk)
        redirect_url = (
            reverse("list_goals", kwargs={"pk": project.pk}) + f"?goal_pk={goal_pk}"
        )

        task_form = CreateTaskForm(request.POST)
        if task_form.is_valid():
            task = task_form.save(commit=False)
            task.goal = goal
            task.save()
            task_form.save_m2m()  # Guarda los recursos seleccionados
            messages.success(request, "Tarea creada correctamente!")
            return redirect(redirect_url)

        messages.error(request, "Error al crear la tarea.")
        return redirect(redirect_url)


class TaskDeleteView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        task_id = request.POST.get("task_id")
        task = get_object_or_404(Task, pk=task_id)
        goal = get_object_or_404(Goal, pk=task.goal.pk)
        project = get_object_or_404(Project, pk=goal.project.pk)

        # url de redirección
        redirect_url = reverse("list_goals", kwargs={"pk": project.pk})
        # si hay un goal_pk en la url, lo añadimos como parámetro
        redirect_url += f"?goal_pk={goal.pk}"

        task.delete()
        messages.success(request, "Tarea eliminada correctamente!")
        return redirect(redirect_url)


class TaskGenerateView(LoginRequiredMixin, View):
    def get(self, request, goal_pk, *args, **kwargs):
        goal = get_object_or_404(Goal, pk=goal_pk)
        project = goal.project
        gg = GeminiGenerator()
        task_data = gg.generate_task(goal, project.description)
        # Crea la tarea en la base de datos
        from .models import Task

        task = Task.objects.create(
            name=task_data["name"],
            description=task_data["description"],
            duration_hours=task_data["duration_hours"],
            goal=goal,
        )
        messages.success(request, "Tarea generada correctamente!")
        return redirect(
            f"{reverse('list_goals', kwargs={'pk': project.pk})}?goal_pk={goal.pk}"
        )


class TaskOverwriteWithAIView(LoginRequiredMixin, View):
    def get(self, request, task_pk, *args, **kwargs):
        task = get_object_or_404(Task, pk=task_pk)
        goal = task.goal
        project = goal.project
        gg = GeminiGenerator()
        # Genera nuevos datos para la tarea usando el contexto actual
        task_data = gg.generate_task(goal, project.description)
        # Sobrescribe los datos de la tarea
        task.name = task_data["name"]
        task.description = task_data["description"]
        task.duration_hours = task_data["duration_hours"]
        task.save()
        messages.success(request, "Tarea sobrescrita con IA correctamente!")
        return redirect(
            f"{reverse('list_goals', kwargs={'pk': project.pk})}?goal_pk={goal.pk}"
        )


# Vistas de Objetivos
class GoalsListView(LoginRequiredMixin, ListView):
    template_name = "view/goalList/main.html"
    context_object_name = "goals"
    paginate_by = 10

    def get_queryset(self):
        self.project = get_object_or_404(Project, pk=self.kwargs["pk"])
        search_query = self.request.GET.get("search", "").strip()
        get_goal_pk = self.request.GET.get("goal_pk")

        if get_goal_pk and get_goal_pk != "":
            queryset = self.project.goal_set.filter(pk=get_goal_pk)
        else:
            queryset = Goal.objects.filter(project=self.project).order_by("order")

        if search_query:
            # Filtrar metas que tienen tareas que coinciden con el término de búsqueda
            queryset = queryset.filter(task__name__icontains=search_query).distinct()

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["project"] = self.project
        context["search_query"] = self.request.GET.get("search", "")

        # Filtrar las tareas de cada meta según el término de búsqueda
        search_query = self.request.GET.get("search", "").strip()
        if search_query:
            for goal in context["goals"]:
                goal.filtered_tasks = goal.task_set.filter(name__icontains=search_query)
        else:
            for goal in context["goals"]:
                goal.filtered_tasks = goal.task_set.all()

        return context


class GoalOrderView(LoginRequiredMixin, View):
    def get(self, request, pk, goalpk, action):
        current_goal = get_object_or_404(Goal, pk=goalpk, project__pk=pk)

        # Obtenemos todos los objetivos ordenados del proyecto
        goals = list(
            Goal.objects.filter(project=current_goal.project).order_by("order")
        )
        current_index = next(
            (i for i, g in enumerate(goals) if g.pk == current_goal.pk), None
        )

        with transaction.atomic():
            if action == "up":
                self._move_up(current_goal, goals, current_index, request)
            elif action == "down":
                self._move_down(current_goal, goals, current_index, request)
            else:
                messages.error(request, "Acción no válida o movimiento imposible.")

        return redirect(reverse("list_goals", kwargs={"pk": pk}))

    def _move_up(self, current_goal, goals, current_index, request):
        if current_index > 0:
            # Intercambiar con el anterior
            previous_goal = goals[current_index - 1]
            current_goal.order, previous_goal.order = (
                previous_goal.order,
                current_goal.order,
            )
            current_goal.save()
            previous_goal.save()
        else:
            messages.error(request, "No se puede subir más.")

    def _move_down(self, current_goal, goals, current_index, request):
        if current_index < len(goals) - 1:
            # Intercambiar con el siguiente
            next_goal = goals[current_index + 1]
            current_goal.order, next_goal.order = next_goal.order, current_goal.order
            current_goal.save()
            next_goal.save()
        else:
            messages.error(request, "No se puede bajar más.")


class GoalUpdateView(LoginRequiredMixin, UpdateView):
    model = Goal
    fields = ["name", "description"]

    def get_object(self):
        goal_id = self.request.POST.get("id")
        project_id = self.kwargs.get("pk")
        return get_object_or_404(Goal, pk=goal_id, project__id=project_id)

    def get_success_url(self):
        return reverse("list_goals", kwargs={"pk": self.kwargs.get("pk")})

    def form_valid(self, form):
        messages.success(
            self.request,
            f"Meta '{form.cleaned_data.get('name')}' actualizada correctamente!",
        )
        return super().form_valid(form)

    def form_invalid(self, form):
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(
                    self.request, f"Error en {form.fields[field].label}: {error}"
                )
        return redirect(self.get_success_url())


class GoalDeleteView(LoginRequiredMixin, View):
    """
    Vista para eliminar una meta (Goal).
    La URL incluye el project_id como pk y el goal_id debe venir por POST.
    """

    def post(self, request, pk, *args, **kwargs):
        try:
            goal_id = request.POST.get("goal_id")

            if not goal_id:
                messages.error(request, "Falta el ID de la meta a eliminar.")
                return redirect("home")

            project = get_object_or_404(Project, id=pk)
            goal = get_object_or_404(Goal, id=goal_id, project=project)

            goal.delete()

            messages.success(request, "¡Meta eliminada correctamente!")
        except Exception as e:
            messages.error(request, f"Error al eliminar la meta: {str(e)}")

        return redirect(reverse("list_goals", kwargs={"pk": pk}))


class GoalCreateView(LoginRequiredMixin, CreateView):
    model = Goal
    fields = ["name", "description"]

    def get_success_url(self):
        return reverse("list_goals", kwargs={"pk": self.kwargs.get("pk")})

    def form_valid(self, form):
        # Asignar automáticamente el proyecto desde la URL
        project = get_object_or_404(Project, pk=self.kwargs.get("pk"))
        form.instance.project = project

        messages.success(
            self.request,
            f"Meta '{form.cleaned_data.get('name', 'nuevo objetivo')}' creado correctamente!",
        )

        return super().form_valid(form)

    def form_invalid(self, form):
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(
                    self.request, f"Error en {form.fields[field].label}: {error}"
                )

        return redirect(reverse("list_goals", kwargs={"pk": self.kwargs.get("pk")}))


class GoalGenerateNewView(LoginRequiredMixin, View):
    def get(self, request, pk, *args, **kwargs):
        project = get_object_or_404(Project, pk=pk)
        gg = GeminiGenerator()
        goal_obj = gg.generate_goal(project, project.description, 0)
        Goal.objects.create(
            name=goal_obj.name, description=goal_obj.description, project=project
        )
        messages.success(request, "Meta generada correctamente con IA.")
        # Mantener los parámetros GET (como page)
        params = request.GET.copy()

        print(params)
        url = reverse("list_goals", kwargs={"pk": pk})
        if params:
            url += "?" + urlencode(params)
        return redirect(url)


class GoalOverwriteWithAIView(LoginRequiredMixin, View):
    def get(self, request, pk, goalpk, *args, **kwargs):
        project = get_object_or_404(Project, pk=pk)
        goal = get_object_or_404(Goal, pk=goalpk, project=project)
        gg = GeminiGenerator()
        goal2 = gg.generate_goal(project, project.description, goalpk)
        goal.name = goal2.name
        goal.description = goal2.description
        goal.save()
        messages.success(request, "Meta sobrescrita correctamente con IA.")
        return redirect(f"{reverse('list_goals', kwargs={'pk': pk})}?goal_pk={goal.pk}")


# Vistas de Proyecto
class ProjectDescriptionAICreateFormView(LoginRequiredMixin, FormView):
    template_name = "view/project_description_ai_form.html"
    form_class = EmprendedorDescripcionForm

    def get_initial(self):
        # Recupera los datos previos de la sesión si existen
        return self.request.session.get("project_ai_data", {})

    def form_valid(self, form):
        data = form.cleaned_data
        self.request.session["project_ai_data"] = data
        return redirect("project_description_ai_edit")


class ProjectDescriptionAIEditView(LoginRequiredMixin, FormView):
    template_name = "view/project_description_ai_edit.html"
    success_url = reverse_lazy("home")

    class EditForm(forms.Form):
        name = forms.CharField(
            label="Nombre del proyecto",
            widget=forms.Textarea(attrs={"style": "height: 60px;"}),
        )
        description = forms.CharField(
            label="Descripción del proyecto",
            widget=forms.Textarea(attrs={"style": "height: 200px;"}),
        )

    form_class = EditForm

    def get_initial(self):
        data = self.request.session.get("project_ai_data")
        if not data:
            return {}
        gg = GeminiGenerator()
        descripcion = gg.generate_project_description(data)
        return {"name": data.get("idea", ""), "description": descripcion}

    def get(self, request, *args, **kwargs):
        # Si no hay datos, redirige al formulario de preguntas
        if not self.request.session.get("project_ai_data"):
            return redirect("project_description_ai_form")
        return super().get(request, *args, **kwargs)

    def form_valid(self, form):
        Project.objects.create(
            name=form.cleaned_data["name"],
            description=form.cleaned_data["description"],
            owner=self.request.user,
            creation_date=timezone.now(),
            last_modified=timezone.now(),
        )
        # Limpia la sesión
        self.request.session.pop("project_ai_data", None)
        messages.success(self.request, "Proyecto creado correctamente con IA.")
        return super().form_valid(form)


class ExportProjectExcelView(LoginRequiredMixin, View):
    def get(self, request, pk):
        project = get_object_or_404(Project, pk=pk, owner=request.user)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Proyecto"

        # Encabezados
        headers = ["Meta", "Tarea", "Duración (h)", "Recurso(s)"]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws[f"{get_column_letter(col)}1"].font = Font(bold=True)

        row_num = 2
        for goal in project.goal_set.all().order_by("order"):
            if goal.task_set.exists():
                for task in goal.task_set.all():
                    recursos = ", ".join([res.name for res in task.resources.all()])
                    ws.append([goal.name, task.name, task.duration_hours, recursos])
                    row_num += 1
            else:
                ws.append([goal.name, "", "", ""])
                row_num += 1

        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="proyecto_{project.id}.xlsx"'
        )
        wb.save(response)
        return response


class ExportProjectCSVView(LoginRequiredMixin, View):
    def get(self, request, pk):
        project = get_object_or_404(Project, pk=pk, owner=request.user)
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            f'attachment; filename="proyecto_{project.id}.csv"'
        )

        writer = csv.writer(response)
        writer.writerow(["Meta", "Tarea", "Duración (h)", "Recurso(s)"])

        for goal in project.goal_set.all().order_by("order"):
            for task in goal.task_set.all():
                recursos = ", ".join([res.name for res in task.resources.all()])
                writer.writerow([goal.name, task.name, task.duration_hours, recursos])
            # Si una meta no tiene tareas, igual la mostramos
            if not goal.task_set.exists():
                writer.writerow([goal.name, "", "", ""])

        return response


class ProjectEditView(LoginRequiredMixin, View):
    template_name = "view/projectDetails/main.html"

    def get(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        form = ProjectEditForm(instance=project)
        context = {"form": form, "project": project}
        return render(request, self.template_name, context)

    def post(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        form = ProjectEditForm(request.POST, instance=project)

        if form.is_valid():
            project = form.save(commit=False)
            project.last_modified = timezone.now()
            project.save()
            messages.success(request, "Proyecto actualizado correctamente")

        context = {"form": form, "project": project}
        return render(request, self.template_name, context)


class DeleteProjectView(LoginRequiredMixin, FormView):
    form_class = ArchiveProjectForm
    success_url = reverse_lazy("home")

    def form_valid(self, form):
        project_id = form.cleaned_data["project_id"]
        project = get_object_or_404(Project, pk=project_id)

        if self.request.user == project.owner:
            project.delete()
            messages.success(self.request, "Proyecto eliminado correctamente")
        else:
            messages.error(
                self.request, "No tienes permiso para eliminar este proyecto"
            )
            return self.form_invalid(form)

        return super().form_valid(form)


class ArchiveProjectView(LoginRequiredMixin, FormView):
    form_class = ArchiveProjectForm
    success_url = reverse_lazy("home")

    def form_valid(self, form):
        project_id = form.cleaned_data["project_id"]

        try:
            project = Project.objects.get(pk=project_id)
            project.is_archived = True
            project.save()

            messages.success(self.request, f"Proyecto archivado correctamente!")

        except Project.DoesNotExist:
            messages.error(self.request, f"Campo es obligatorio.")
            pass

        return super().form_valid(form)


class ProjectCreateView(LoginRequiredMixin, CreateView):
    model = Project
    fields = ["name", "description", "is_archived"]
    success_url = reverse_lazy("home")

    def form_valid(self, form):
        user = self.request.user
        # Verifica si el usuario es básico y ya tiene un proyecto
        if hasattr(user, "profile") and user.profile.account_type == "basic":
            if Project.objects.filter(owner=user).count() >= 1:
                messages.error(
                    self.request,
                    "Con una cuenta básica solo puedes crear un proyecto. Actualiza a Premium para crear más."
                )
                return redirect(self.success_url)
        form.instance.owner = user
        form.instance.creation_date = timezone.now()
        form.instance.last_modified = timezone.now()

        project_name = form.cleaned_data.get("name", "nuevo proyecto")
        messages.success(
            self.request, f"Proyecto '{project_name}' creado correctamente!"
        )

        return super().form_valid(form)

    def form_invalid(self, form):
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(
                    self.request, f"Campo '{form.fields[field].label}' es obligatorio."
                )

        return redirect(self.success_url)


class ProjectListView(LoginRequiredMixin, ListView):
    model = Project
    template_name = "view/proyectList/main.html"
    context_object_name = "projects"
    ordering = ["-last_modified"]
    paginate_by = 10

    def get_queryset(self):
        user = self.request.user
        # Mostrar proyectos donde soy dueño o colaborador
        queryset = Project.objects.filter(
            Q(owner=user) | Q(collaborator__user=user)
        ).distinct()
        search_query = self.request.GET.get("search", "").strip()

        if self.request.GET.get("filter") == "archived":
            queryset = queryset.filter(is_archived=True)

        if self.request.GET.get("filter") == "active":
            queryset = queryset.filter(is_archived=False)

        # Aplicar búsqueda si hay un término de búsqueda
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query)
                | Q(owner__username__icontains=search_query)
            )

        return queryset.select_related("owner")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        query_params = self.request.GET.copy()
        if "page" in query_params:
            del query_params["page"]
        context["current_query"] = urlencode(query_params)
        context["search_query"] = self.request.GET.get("search", "")
        context["current_filter"] = self.request.GET.get("filter", "")
        return context